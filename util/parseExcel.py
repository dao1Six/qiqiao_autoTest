# encoding = utf-8
from openpyxl import load_workbook


class ParseExcel(object):

    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        self.ws = self.wb.get_sheet_by_name(sheetName)

    # 获取表格的总行数和总列数
    def getRowsClosNum( self ):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue( self, row, column ):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues( self, column ):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata


    # 设置某个单元格的值
    def setCellValue( self, row, colunm, cellvalue ):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    # 获取某行所有值
    def getRowValues( self, row ):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    #获取一批行数据返回列表
    def getRowsValues(self,start,end):
        dataList = []
        for i in range(start,end+1):
            print(i)
            rowValue = self.getRowValues(i)
            dataList.append(rowValue)
        return dataList


    def getSheetValue( self ):
        #获取sheetd的总行数
        rows = self.getRowsClosNum()[0]
        dataList = []
        for row in range(1,rows+1):
            rowValue = self.getRowValues(row)
            dataList.append(rowValue)
        return dataList
